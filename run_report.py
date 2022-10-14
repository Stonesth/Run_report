from Tools import tools_v000 as tools
from Jira import jira as j
import os
from os.path import dirname
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

font = Font(name='ING Me', size=12, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='00FF6600')



# -10 for the name of this project Run_report
save_path = os.path.dirname(os.path.abspath("__file__"))
propertiesFolder_path = save_path + "/"+ "Properties"

# Example of used
j.jira = tools.readProperty(propertiesFolder_path, 'Run_report', 'jira=')

tools.openBrowserChrome()

# Open the JIRA with the Run that you want to check
j.connectToJira(j.jira)

# Recovered all the Issue in epic
tools.waitLoadingPageByID2(10, "ghx-issues-in-epic-table")

rows = tools.driver.find_elements_by_xpath('//*[@id="ghx-issues-in-epic-table"]/tbody/tr')
temp = [] # Temproary list
for row in rows:
   temp.append(row.get_attribute('data-issuekey'))

tools.closeBrowserChrome()

# create a workbook as .xlsx file
def create_workbook(path):
   workbook = Workbook()
   workbook.save(path) 
if __name__ == "__main__": 
   create_workbook("file.xlsx") 
print("File Created Successfully")

wb = load_workbook('file.xlsx') 
# Sheet is the SheetName where the data has to be entered 
sheet = wb['Sheet'] 
# Enter into 1st row and Ath column 
sheet['A1'] = 'CINS-IT4IT - RUN ACTION LIST - V1 2022' 
# Similarly you can enter in the below shown fashion 
sheet.cell(row=3, column=1).value = 'SPRINT' 
sheet.cell(row=3, column=2).value = 'WHO' 
sheet.cell(row=3, column=3).value = 'TYPE' 
sheet.cell(row=3, column=4).value = 'INCIDENT REF' 
sheet.cell(row=3, column=5).value = 'JIRA REF' 
sheet.cell(row=3, column=6).value = 'DIRECT IMPACT BROKER' 
sheet.cell(row=3, column=7).value = 'APP'
sheet.cell(row=3, column=8).value = 'CINS DOMAINE'
sheet.cell(row=3, column=9).value = 'DATA TYPE'
sheet.cell(row=3, column=10).value = 'COMMENT'
sheet.cell(row=3, column=11).value = 'ID'

# Place font
sheet['A1'].font = font
sheet.cell(row=3, column=1).font = font
sheet.cell(row=3, column=2).font = font
sheet.cell(row=3, column=3).font = font
sheet.cell(row=3, column=4).font = font
sheet.cell(row=3, column=5).font = font
sheet.cell(row=3, column=6).font = font
sheet.cell(row=3, column=7).font = font
sheet.cell(row=3, column=8).font = font
sheet.cell(row=3, column=9).font = font
sheet.cell(row=3, column=10).font = font
sheet.cell(row=3, column=11).font = font

i = 0
for jira in temp:
   i = i + 1
   print(jira)
   j.jira = jira

   tools.openBrowserChrome()
   j.connectToJira(j.jira)

   # Collect info in the page
   # Sprint
   try :
      tools.waitLoadingPageByID2(10, 'customfield_10007-val')
      sprint = tools.driver.find_element_by_xpath('//*[@id="customfield_10007-val"]').text
   except :
      print("No sprint for this one : ")
      sprint = 'No Sprint'
      pass
   
   # Assignee
   tools.waitLoadingPageByID2(10, 'assignee-val')
   assignee = tools.driver.find_element_by_xpath('//*[@id="assignee-val"]').text
   
   # Type = RUN
   type = "RUN"
   
   # INCIDENT REF
   incident_ref = ''   
   
   # JIRA REF
   
   # DIRECT IMPACT BROKER
   direct_impact_broker = ''
   
   # APP
   application = ''

   # CINS DOMAIN
   cins_domain = ''
   
   # DATA TYPE
   data_type = ''
   
   # COMMENT
   try:
      tools.waitLoadingPageByID2(10, 'summary-val')
      comment = tools.driver.find_element_by_xpath('//*[@id="summary-val"]').text
      print('Comment = ' + comment)
   except UnicodeEncodeError as ex :
      try:
         print("UnicodeEncodeError : ")
         comment = tools.driver.find_element_by_xpath('//*[@id="summary-val"]').text.encode('utf-8')
         pass
      except:
         print("Error")
         comment = ''
         pass

   # Split elements 
   # https://www.w3schools.com/python/ref_string_split.asp
   x = sprint.split(", ")
   # last element 
   # https://appdividend.com/2022/06/23/how-to-get-last-element-of-a-list-in-python/#:~:text=To%20get%20the%20last%20element,to%20get%20the%20last%20element.
   print(x[-1])
   
       
   print('Sprint = ' + sprint)
   print('Assignee = ' + assignee)
   print('Type = ' + type)
   print('Comment = ' + comment)
   
   # Sprint
   sheet.cell(row=3+i, column=1).value = sprint
   # Assignee
   sheet.cell(row=3+i, column=2).value = assignee 
   # Type = RUN
   sheet.cell(row=3+i, column=3).value = type 
   # INCIDENT REF
   sheet.cell(row=3+i, column=4).value = '' 
   # JIRA REF
   sheet.cell(row=3+i, column=5).value = jira
   # DIRECT IMPACT BROKER
   sheet.cell(row=3+i, column=6).value = '' 
   # APP
   sheet.cell(row=3+i, column=7).value = ''
   # CINS DOMAIN
   sheet.cell(row=3+i, column=8).value = ''
   # DATA TYPE
   sheet.cell(row=3+i, column=9).value = ''
   # COMMENT
   sheet.cell(row=3+i, column=10).value = comment
   # ID
   sheet.cell(row=3+i, column=11).value = i   
   
   tools.closeBrowserChrome()
   

wb.save('file.xlsx')
   # # Get all the columns for each row. 
   # # cols = row.find_elements_by_xpath("./*")
   # cols = row.find_elements_by_xpath("./*[name()='th' or name()='td']")
   # temp = [] # Temproary list
   # for col in cols:
   #    temp.append(col.text)
   # print(temp)   
   
   
   # print(row.get_attribute('data-issuekey'))
   # # Get all the columns for each row. 
   # # cols = row.find_elements_by_xpath("./*")
   # cols = row.find_elements_by_xpath("//td[contains(@class,'nav ghx-summary')]")
   # for col in cols:
   #    print(col.text)
   
   # cols = row.find_elements_by_xpath("//td[contains(@class,'nav assignee')]")
   # for col in cols:
   #    print(col.text)
        
   
   # print(row.text)
   # for col in row.find_elements_by_xpath("//td"):
   #    print(col.text)
          
# tools.closeBrowserChrome()
